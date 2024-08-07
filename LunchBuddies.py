import sys
import os
import openpyxl
import itertools
import random


def usage() -> None:
    print(r"""
    Improper command line usage!
   
    PROPER COMMAND LINE USAGE:
        py.exe .\LunchBuddies.py excel-file-path numWeeks

    EXAMPLE USAGE:
        py.exe .\LunchBuddies.py "LunchBuddiesInfo.xlsx" 3
        py.exe .\LunchBuddies.py "C:\Users\User.Name\Desktop\LunchBuddiesInfo.xlsx" 13

    BAD USAGE:
        py.exe .\LunchBuddies.py 
        py.exe .\LunchBuddies.py 3 
        py.exe .\LunchBuddies.py "LunchBuddiesInfo.xlsx" 
""")


# Generates and returns all possible pairings
def generateUniquePairings(execs, generals):
    allPairings = list(itertools.product(execs, generals))
    random.shuffle(allPairings)  
    return allPairings


def writePairingsToExcel(wb, numWeeks, execs, generals):
    # Calculates the total possible unique pairings
    maxUniquePairings = min(len(execs), len(generals))

    if numWeeks > maxUniquePairings:
        print(f"Cannot create {numWeeks} weeks of unique pairings with the given lists.")
        sys.exit(1)

    allPairings = generateUniquePairings(execs, generals)

    # Start creating sheets and assigning pairings
    for week in range(numWeeks):
        # Create a new sheet for each week
        sheetName = f"Week {week + 1}"
        sheet = wb.create_sheet(sheetName)
        usedPairings = set()

        # Assign unique pairings for this week
        row = 1
        for pairing in allPairings:
            if pairing not in usedPairings:
                exec_member, general_member = pairing
                sheet[f"A{row}"] = exec_member
                sheet[f"B{row}"] = general_member
                row += 1
                usedPairings.add(pairing)

                # Once we have enough pairings for this week, break out of the loop
                if len(usedPairings) == maxUniquePairings:
                    break

        # Handle internal pairing within the larger group if necessary
        remaining_members = (generals if len(generals) > len(execs) else execs).copy()
        random.shuffle(remaining_members)
        
        while len(remaining_members) > 1:
            member1 = remaining_members.pop()
            member2 = remaining_members.pop()
            sheet[f"A{row}"] = member1
            sheet[f"B{row}"] = member2
            row += 1

        # If there's an odd one out, they will not be paired this week
        if remaining_members:
            sheet[f"A{row}"] = remaining_members[0]
            sheet[f"B{row}"] = "No Pairing"

        # Remove used pairings for subsequent weeks
        allPairings = [pair for pair in allPairings if pair not in usedPairings]


def openExcelFile(wb) -> dict:
    sheet = wb.active
    names = {}

    for column in sheet.iter_cols(): 
        column_name = column[0].value 
        names[column_name] = []
        for i, cell in enumerate(column): 
            if i == 0 or cell.value is None:
                continue
            names[column_name].append(cell.value) 
    return names


if __name__ == "__main__":
    if len(sys.argv) != 3:
        usage()
        sys.exit(1)

    excelFile = sys.argv[1]
    numWeeks = sys.argv[2]

    try:
        numWeeks = int(sys.argv[2])
    except ValueError:
        print("Please provide a valid integer for the number of weeks.")
        sys.exit(1)

    if not os.path.isfile(excelFile):
        print(f"The file {excelFile} does not exist.")
        sys.exit(1)

    wb = openpyxl.load_workbook(excelFile)

    # dictionary of executive member names and general member names
    names = openExcelFile(wb)

    columnNames = list(names.keys())
    writePairingsToExcel(wb, numWeeks, names[columnNames[0]], names[columnNames[1]])
    wb.save(excelFile)

    print(f"Open {excelFile} to see the lunch buddies!")

    



